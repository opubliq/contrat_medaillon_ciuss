"""
generate_livrable.py

Lit 06_bottin_complet.csv et génère le livrable xlsx
liste_organismes/data/livrable_liste_organismes.xlsx avec 4 sheets :
- Local (local + hochelaga-maisonneuve)
- Montréal
- Grand Montréal
- Provincial

Colonnes : Organisme, Description, Secteur, Courriel, Téléphone, Site web
Filtre : exclu == 'non' seulement.
"""

import csv
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = SCRIPT_DIR.parent / "data"
INPUT_FILE = DATA_DIR / "06_bottin_complet.csv"
OUTPUT_FILE = DATA_DIR / "livrable_liste_organismes.xlsx"

# Mapping secteur → libellé lisible
SECTEUR_LABELS = {
    "sante": "Santé",
    "defense_droits": "Défense des droits",
    "soutien_vulnerabilite": "Soutien aux personnes vulnérables",
    "developpement_enfants_familles": "Développement enfants et familles",
    "developpement_communautaire": "Développement communautaire",
    "education_formation": "Éducation et formation",
    "emploi": "Emploi",
    "autres_missions_sociales": "Autres missions sociales",
    "logement": "Logement",
    "immigration": "Immigration et intégration",
    "violence_victimes": "Violence et victimes",
    "dependances": "Dépendances",
    "alimentation": "Alimentation",
    "transport": "Transport",
    "loisirs_culture": "Loisirs et culture",
}

COLUMNS = ["Organisme", "Description", "Secteur", "Courriel", "Téléphone", "Site web"]

SHEETS = [
    ("Local", ["local", "hochelaga-maisonneuve"]),
    ("Montréal", ["montreal"]),
    ("Grand Montréal", ["grand_montreal"]),
    ("Provincial", ["provincial"]),
]

# Styles
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
ROW_FONT = Font(name="Calibri", size=10)
ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
COL_WIDTHS = [40, 55, 30, 30, 18, 30]


def load_data() -> list[dict]:
    with open(INPUT_FILE, encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    return [r for r in rows if r["exclu"].strip() == "non"]


def secteur_label(code: str) -> str:
    code = code.strip()
    return SECTEUR_LABELS.get(code, code.replace("_", " ").title())


def build_row(r: dict) -> list:
    secteur = secteur_label(r["secteur"])
    if r.get("secteur_secondaire", "").strip():
        secteur += f" / {secteur_label(r['secteur_secondaire'])}"
    return [
        r["nom"].strip(),
        r["description"].strip(),
        secteur,
        r["courriel"].strip(),
        r["telephone"].strip(),
        r["site_web"].strip(),
    ]


def write_sheet(wb: openpyxl.Workbook, sheet_name: str, categories: list[str], data: list[dict]):
    rows = [r for r in data if r["categorie_territoire"].strip() in categories]
    rows.sort(key=lambda r: r["nom"].lower())

    ws = wb.create_sheet(title=sheet_name)

    # Header
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx, r in enumerate(rows, start=2):
        values = build_row(r)
        alt = (row_idx % 2 == 0)
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = ROW_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if alt:
                cell.fill = ALT_FILL
        ws.row_dimensions[row_idx].height = 30

    # Column widths
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    print(f"  Sheet '{sheet_name}': {len(rows)} organismes")


def main():
    data = load_data()
    print(f"Total organismes non-exclus: {len(data)}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for sheet_name, categories in SHEETS:
        write_sheet(wb, sheet_name, categories, data)

    wb.save(OUTPUT_FILE)
    print(f"\n✓ Livrable généré : {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
